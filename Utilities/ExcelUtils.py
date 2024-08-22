import openpyxl


def getRowCount(file, Sheet1):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[Sheet1]
    return sheet.max_row


def getColumnCount(file, Sheet1):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[Sheet1]
    return sheet.max_Column


def readData(file, Sheet1, rownum, columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[Sheet1]
    return sheet.cell(row=rownum, column=columnnum).value


def writeData(file, Sheet1, rownum, columnnum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[Sheet1]
    sheet.cell(row=rownum, column=columnnum).value = data
    workbook.save(file)
